#!/usr/bin/env python3
"""Import candidates from Excel file into Supabase candidates table."""

import re
import openpyxl
from supabase import create_client

SUPABASE_URL = "https://psrsosfjteeovtmszwgu.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InBzcnNvc2ZqdGVlb3Z0bXN6d2d1Iiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzE1MzI2MDAsImV4cCI6MjA4NzEwODYwMH0.15QS6GQ2cEWc-a1OVvzT1DlrExbdWGoRdEnYZ-ypgZs"
EXCEL_PATH = "/Users/nikitaguzenko/Downloads/divine ent driver tracking.xlsx"

sb = create_client(SUPABASE_URL, SUPABASE_KEY)


def normalize_phone(raw):
    """Strip non-digits, remove leading 1 if 11 digits, format XXX-XXX-XXXX. Return None if invalid."""
    if not raw:
        return None
    # Handle float values like 5303083162.0
    val = str(raw)
    if val.endswith(".0"):
        val = val[:-2]
    digits = re.sub(r"\D", "", val)
    if len(digits) == 11 and digits.startswith("1"):
        digits = digits[1:]
    if len(digits) != 10:
        return None
    return f"{digits[:3]}-{digits[3:6]}-{digits[6:]}"


def normalize_status(raw):
    """Map raw status to canonical status."""
    if not raw or str(raw).strip() == "":
        return "New"
    s = str(raw).strip()
    sl = s.lower()
    # Phone number as status
    if re.match(r"^[\d\-\(\)\s\+\.]+$", s) and len(re.sub(r"\D", "", s)) >= 7:
        return "New"
    if sl in ("no answer", "no answer 2x", "no answer 2"):
        return "No Answer"
    if sl == "not qualified":
        return "Not Qualified"
    if sl in ("potential driver", "potential"):
        return "New"
    if sl in ("sent link",):
        return "New"
    return s


def split_name(full_name):
    """Split full name into first_name and last_name."""
    if not full_name:
        return None, None
    parts = str(full_name).strip().split(None, 1)
    first = parts[0] if parts else None
    last = parts[1] if len(parts) > 1 else None
    return first, last


def fetch_existing_phones():
    """Fetch all existing phone numbers from Supabase, return set of last-10-digit strings."""
    phones = set()
    offset = 0
    batch = 1000
    while True:
        resp = sb.table("candidates").select("phone").range(offset, offset + batch - 1).execute()
        if not resp.data:
            break
        for row in resp.data:
            if row.get("phone"):
                digits = re.sub(r"\D", "", str(row["phone"]))
                if len(digits) >= 10:
                    phones.add(digits[-10:])
        if len(resp.data) < batch:
            break
        offset += batch
    return phones


def process_standard_sheet(ws, sheet_label, existing_phones):
    """Process Sheet1 or Sheet3 (standard layout):
    col[0]=Name, col[1]=Phone, col[2]=Email, col[3]=Source, col[4]=Status,
    col[8]=Notes, col[9]=Experience
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], 0, 0

    data_rows = rows[1:]  # skip header
    to_insert = []
    skipped_invalid = 0
    skipped_dup = 0

    for row in data_rows:
        if not row or len(row) < 2:
            skipped_invalid += 1
            continue

        name_raw = row[0]
        phone_raw = row[1]
        email = str(row[2]).strip() if len(row) > 2 and row[2] and str(row[2]).strip() else None
        source_raw = str(row[3]).strip() if len(row) > 3 and row[3] else ""
        status_raw = row[4] if len(row) > 4 else None
        notes = str(row[8]).strip() if len(row) > 8 and row[8] else None
        experience = str(row[9]).strip() if len(row) > 9 and row[9] else None

        if not name_raw or str(name_raw).strip().lower() in ("none", ""):
            skipped_invalid += 1
            continue

        phone = normalize_phone(phone_raw)
        if not phone:
            skipped_invalid += 1
            continue

        digits10 = re.sub(r"\D", "", phone)[-10:]
        if digits10 in existing_phones:
            skipped_dup += 1
            continue

        existing_phones.add(digits10)
        first_name, last_name = split_name(name_raw)
        status = normalize_status(status_raw)
        source = "BAZAR.CLUB" if "bazar" in source_raw.lower() else f"Excel/{sheet_label}"

        record = {
            "first_name": first_name,
            "phone": phone,
            "source": source,
            "status": status,
        }
        if last_name:
            record["last_name"] = last_name
        if email:
            record["email"] = email
        if notes:
            record["notes"] = notes
        if experience:
            record["experience"] = experience

        to_insert.append(record)

    return to_insert, skipped_invalid, skipped_dup


def process_bazar_sheet(ws, existing_phones):
    """Process 'CVs from BAZAR.CLUB' sheet (shifted columns):
    col[0]=None, col[1]=Name, col[2]=Phone, col[3]=Source, col[4]=Status
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], 0, 0

    data_rows = rows[1:]  # skip header
    to_insert = []
    skipped_invalid = 0
    skipped_dup = 0

    for row in data_rows:
        if not row or len(row) < 5:
            skipped_invalid += 1
            continue

        # Shifted layout
        name_raw = row[1]
        phone_raw = row[2]
        status_raw = row[4] if len(row) > 4 else None

        if not name_raw or str(name_raw).strip().lower() in ("none", ""):
            skipped_invalid += 1
            continue

        phone = normalize_phone(phone_raw)
        if not phone:
            skipped_invalid += 1
            continue

        digits10 = re.sub(r"\D", "", phone)[-10:]
        if digits10 in existing_phones:
            skipped_dup += 1
            continue

        existing_phones.add(digits10)
        first_name, last_name = split_name(name_raw)
        status = normalize_status(status_raw)

        record = {
            "first_name": first_name,
            "phone": phone,
            "source": "BAZAR.CLUB",
            "status": status,
        }
        if last_name:
            record["last_name"] = last_name

        to_insert.append(record)

    return to_insert, skipped_invalid, skipped_dup


def batch_insert(records, label):
    """Insert records in batches of 50."""
    inserted = 0
    for i in range(0, len(records), 50):
        batch = records[i:i+50]
        resp = sb.table("candidates").insert(batch).execute()
        inserted += len(resp.data) if resp.data else 0
    print(f"  [{label}] Inserted: {inserted}")
    return inserted


def main():
    print("Loading Excel file...")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    print(f"Sheets found: {wb.sheetnames}")

    print("\nFetching existing phones from Supabase...")
    existing_phones = fetch_existing_phones()
    print(f"Existing candidates in DB: {len(existing_phones)}")

    total_inserted = 0
    total_skipped_invalid = 0
    total_skipped_dup = 0

    # Sheet1
    print("\n--- Processing Sheet1 ---")
    ws1 = wb["Sheet1"]
    records, inv, dup = process_standard_sheet(ws1, "Sheet1", existing_phones)
    print(f"  Valid records to insert: {len(records)}, Skipped (invalid): {inv}, Skipped (dup): {dup}")
    if records:
        total_inserted += batch_insert(records, "Sheet1")
    total_skipped_invalid += inv
    total_skipped_dup += dup

    # Sheet3
    print("\n--- Processing Sheet3 ---")
    ws3 = wb["Sheet3"]
    records, inv, dup = process_standard_sheet(ws3, "Sheet3", existing_phones)
    print(f"  Valid records to insert: {len(records)}, Skipped (invalid): {inv}, Skipped (dup): {dup}")
    if records:
        total_inserted += batch_insert(records, "Sheet3")
    total_skipped_invalid += inv
    total_skipped_dup += dup

    # CVs from BAZAR.CLUB
    print("\n--- Processing 'CVs from BAZAR.CLUB' ---")
    ws_bazar = wb["CVs from BAZAR.CLUB"]
    records, inv, dup = process_bazar_sheet(ws_bazar, existing_phones)
    print(f"  Valid records to insert: {len(records)}, Skipped (invalid): {inv}, Skipped (dup): {dup}")
    if records:
        total_inserted += batch_insert(records, "BAZAR.CLUB")
    total_skipped_invalid += inv
    total_skipped_dup += dup

    wb.close()

    print(f"\n{'='*50}")
    print(f"TOTAL INSERTED: {total_inserted}")
    print(f"TOTAL SKIPPED (invalid/no phone): {total_skipped_invalid}")
    print(f"TOTAL SKIPPED (duplicate): {total_skipped_dup}")
    print(f"{'='*50}")

    # Verify BAZAR.CLUB
    print("\nVerification - BAZAR.CLUB candidates in DB:")
    resp = sb.table("candidates").select("id, first_name, last_name, phone, status").eq("source", "BAZAR.CLUB").execute()
    print(f"  Count: {len(resp.data)}")
    for r in resp.data[:5]:
        print(f"    {r['id']}: {r['first_name']} {r.get('last_name') or ''} | {r['phone']} | {r['status']}")
    if len(resp.data) > 5:
        print(f"    ... and {len(resp.data) - 5} more")


if __name__ == "__main__":
    main()
